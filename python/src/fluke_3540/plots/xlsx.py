"""Generate a chartable XLSX workbook from a session CSV.

Sheets: Summary, Load Profile, Power Factor, Harmonics, Voltage & Current, Data.
Asset / team / instrument metadata is read from the session's config dict
(typically the *-config.json next to trend.bin) — no hardcoded names.
"""
from __future__ import annotations

import csv
import datetime as dt
from pathlib import Path
from typing import Mapping

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Slim column list — only what makes sense to plot or eyeball.
COLS_TO_KEEP: list[tuple[str, str]] = [
    ("timestamp_utc",      "Timestamp (UTC)"),
    ("V_LN_a_avg_V",       "V_LN_a (V)"),
    ("V_LN_b_avg_V",       "V_LN_b (V)"),
    ("V_LN_c_avg_V",       "V_LN_c (V)"),
    ("I_a_avg_A",          "I_a (A)"),
    ("I_b_avg_A",          "I_b (A)"),
    ("I_c_avg_A",          "I_c (A)"),
    ("P_total_avg_W",      "P_total (W)"),
    ("S_total_avg_VA",     "S_total (VA)"),
    ("Q_total_avg_VAR",    "Q_total (VAR)"),
    ("PF_total_avg",       "PF (true)"),
    ("DPF_total_avg",      "DPF (displacement)"),
    ("freq_avg_Hz",        "Frequency (Hz)"),
    ("V_THD_pct_a_avg",    "V_THD_a (%)"),
    ("V_THD_pct_b_avg",    "V_THD_b (%)"),
    ("V_THD_pct_c_avg",    "V_THD_c (%)"),
    ("I_THD_pct_a_avg",    "I_THD_a (%)"),
    ("I_THD_pct_b_avg",    "I_THD_b (%)"),
    ("I_THD_pct_c_avg",    "I_THD_c (%)"),
    ("Wh_total",           "Wh_total (per row)"),
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")


def _make_scatter(ws_data, display_names: list[str], data_last_row: int,
                  chart_title: str, y_axis: str, series_cols: list[str],
                  height: int = 11, width: int = 30) -> ScatterChart:
    chart = ScatterChart()
    chart.title = chart_title
    chart.style = 2
    chart.x_axis.title = "Time (UTC)"
    chart.y_axis.title = y_axis
    chart.x_axis.number_format = "mm/dd hh:mm"
    chart.x_axis.majorTimeUnit = "hours"
    chart.height = height
    chart.width = width
    chart.legend.position = "b"
    x_ref = Reference(ws_data, min_col=1, min_row=2, max_row=data_last_row)
    for name in series_cols:
        if name not in display_names:
            continue
        col_idx = display_names.index(name) + 1
        y_ref = Reference(ws_data, min_col=col_idx, min_row=2, max_row=data_last_row)
        chart.series.append(Series(y_ref, x_ref, title_from_data=False, title=name))
    return chart


def _build_summary(csv_path: Path) -> dict[str, float | int]:
    """Compute energy/peak/time-class totals from the per-second CSV."""
    wh_sum = wh_fwd = wh_rev = 0.0
    p_peak_pos = 0.0
    p_peak_neg = 0.0
    sec_import = sec_export = sec_idle = 0
    i_peak = 0.0
    rows = 0
    with csv_path.open(newline="") as fh:
        reader = csv.DictReader(fh)
        for r in reader:
            try:
                wh = float(r["Wh_total"])
                p = float(r["P_total_avg_W"])
                ia = float(r["I_a_avg_A"])
                ib = float(r["I_b_avg_A"])
                ic = float(r["I_c_avg_A"])
            except (ValueError, KeyError):
                continue
            wh_sum += wh
            if wh > 0:
                wh_fwd += wh
            if wh < 0:
                wh_rev += wh
            if p > p_peak_pos:
                p_peak_pos = p
            if p < p_peak_neg:
                p_peak_neg = p
            if p > 10:
                sec_import += 1
            elif p < -10:
                sec_export += 1
            else:
                sec_idle += 1
            i_peak = max(i_peak, ia, ib, ic)
            rows += 1
    return dict(
        wh_sum=wh_sum, wh_fwd=wh_fwd, wh_rev=wh_rev,
        p_peak_pos=p_peak_pos, p_peak_neg=p_peak_neg, i_peak=i_peak,
        sec_import=sec_import, sec_export=sec_export, sec_idle=sec_idle,
        rows=rows,
    )


def _add_stats_sheet(wb, stats: Mapping) -> None:
    """Add a Statistics sheet from a whole_session_stats() dict."""
    ws = wb.create_sheet("Statistics")
    ws["A1"] = "Whole-session statistics"
    ws["A1"].font = Font(size=14, bold=True)
    headers = ["Channel", "Unit", "Count", "Min", "p1", "p5", "Median",
               "Mean", "p95", "p99", "Max", "Stdev"]
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for name, d in stats.items():
        if name.startswith("_"):
            continue
        ws.append([name, d.get("unit", ""), d["count"],
                   round(d["min"], 3), round(d["p1"], 3), round(d["p5"], 3),
                   round(d["median"], 3), round(d["mean"], 3), round(d["p95"], 3),
                   round(d["p99"], 3), round(d["max"], 3), round(d["stdev"], 3)])
    th = stats.get("_thresholds")
    if th:
        ws.append([])
        ws.append(["Time under-voltage (< %.0f V)" % th["undervoltage_v"],
                   "", f"{th['sec_undervoltage']:,} s",
                   f"{th['pct_undervoltage']:.2f}%"])
        ws.append(["Time over-current (> %.0f A)" % th["overcurrent_a"],
                   "", f"{th['sec_overcurrent']:,} s",
                   f"{th['pct_overcurrent']:.2f}%"])
    ws.column_dimensions["A"].width = 22


def _add_tod_sheet(wb, tod_rows) -> None:
    """Add a Time-of-Day sheet (diurnal avg/min/max envelope) with a chart."""
    ws = wb.create_sheet("Time of Day")
    ws["A1"] = "Time-of-day (diurnal) profile"
    ws["A1"].font = Font(size=14, bold=True)
    header = ["Bin", "n", "n_days", "P avg (kW)", "P min (kW)", "P max (kW)",
              "V avg (V)", "I avg (A)"]
    ws.append([])
    ws.append(header)
    for cell in ws[3]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    first_data = 4
    for r in tod_rows:
        ws.append([r["bin"], r["n"], r["n_days"],
                   round(r["p_avg_kW"], 2), round(r["p_min_kW"], 2),
                   round(r["p_max_kW"], 2), round(r["v_avg_V"], 1),
                   round(r["i_avg_A"], 1)])
    last_data = first_data + len(tod_rows) - 1
    if tod_rows:
        from openpyxl.chart import LineChart, Reference
        chart = LineChart()
        chart.title = "Power by time of day (avg / min / max)"
        chart.y_axis.title = "kW"
        chart.x_axis.title = "Time of day"
        chart.height = 10
        chart.width = 28
        data = Reference(ws, min_col=4, max_col=6, min_row=3, max_row=last_data)
        cats = Reference(ws, min_col=1, min_row=first_data, max_row=last_data)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "J3")
    ws.column_dimensions["A"].width = 10


def write_xlsx(
    csv_path: Path, output_path: Path,
    config: Mapping[str, str | None] | None = None,
    *, csv_per_second_path: Path | None = None,
    stats: Mapping | None = None,
    tod_rows=None,
    narrative: str | None = None,
) -> Path:
    """Build the chartable XLSX.

    csv_path is read for the Data sheet (typically a downsampled CSV).
    csv_per_second_path (defaulting to csv_path) is read for accurate Summary totals.
    config provides asset/team/instrument metadata for the Summary sheet
    (omitted if config is None or empty).
    stats (from whole_session_stats) adds a Statistics sheet; tod_rows (from
    time_of_day_profile) adds a Time-of-Day sheet — both make the report
    self-contained without gnuplot.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary_csv = csv_per_second_path or csv_path
    summary = _build_summary(summary_csv)

    # Load source CSV into memory once
    with csv_path.open(newline="") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
        src_field_names = reader.fieldnames or []

    keep = [(src, disp) for src, disp in COLS_TO_KEEP if src in src_field_names]
    display_names = [disp for _, disp in keep]

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"
    ws_data.append(display_names)
    for cell in ws_data[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        out_row = []
        for src, _ in keep:
            v = r.get(src, "")
            if src == "timestamp_utc":
                try:
                    out_row.append(dt.datetime.fromisoformat(v).replace(tzinfo=None))
                except ValueError:
                    out_row.append(v)
            elif v == "":
                out_row.append(None)
            else:
                try:
                    out_row.append(float(v))
                except ValueError:
                    out_row.append(v)
        ws_data.append(out_row)

    ws_data.freeze_panes = "B2"
    for cell in ws_data["A"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm"
    for col_idx, name in enumerate(display_names, start=1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = max(12, len(name) + 2)

    data_last_row = len(rows) + 1

    # Derived kW/kVA/kVAR columns
    def append_derived(header: str, source_col: str, factor: float) -> None:
        if source_col not in display_names:
            return
        src_idx = display_names.index(source_col) + 1
        new_col_idx = len(display_names) + 1
        cell = ws_data.cell(row=1, column=new_col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        for r in range(2, data_last_row + 1):
            raw = ws_data.cell(row=r, column=src_idx).value
            ws_data.cell(row=r, column=new_col_idx,
                         value=(raw * factor) if isinstance(raw, (int, float)) else None)
        ws_data.column_dimensions[get_column_letter(new_col_idx)].width = max(12, len(header) + 2)
        display_names.append(header)

    append_derived("P_total (kW)",   "P_total (W)",   0.001)
    append_derived("S_total (kVA)",  "S_total (VA)",  0.001)
    append_derived("Q_total (kVAR)", "Q_total (VAR)", 0.001)

    def add_two_chart_sheet(title, top_t, top_y, top_cols, bot_t, bot_y, bot_cols):
        sheet = wb.create_sheet(title)
        sheet["A1"] = title
        sheet["A1"].font = Font(size=14, bold=True)
        sheet["A2"] = "(charts read live from the Data sheet)"
        sheet["A2"].font = Font(italic=True, color="666666")
        sheet.add_chart(_make_scatter(ws_data, display_names, data_last_row,
                                      top_t, top_y, top_cols), "A4")
        sheet.add_chart(_make_scatter(ws_data, display_names, data_last_row,
                                      bot_t, bot_y, bot_cols), "A26")

    def add_single_chart_sheet(title, chart_t, y, cols):
        sheet = wb.create_sheet(title)
        sheet["A1"] = title
        sheet["A1"].font = Font(size=14, bold=True)
        sheet["A2"] = "(chart reads live from the Data sheet)"
        sheet["A2"].font = Font(italic=True, color="666666")
        sheet.add_chart(_make_scatter(ws_data, display_names, data_last_row,
                                      chart_t, y, cols, height=16), "A4")

    add_two_chart_sheet(
        "Load Profile",
        "Real (active) power P over time", "P (kW)  — negative = exporting",
        ["P_total (kW)"],
        "Apparent S and Reactive Q over time", "kVA / kVAR",
        ["S_total (kVA)", "Q_total (kVAR)"],
    )
    add_single_chart_sheet(
        "Power Factor", "True PF vs Displacement PF",
        "Power factor (-1 to +1)", ["PF (true)", "DPF (displacement)"],
    )
    add_two_chart_sheet(
        "Harmonics",
        "Voltage THD per phase", "V_THD (%)",
        ["V_THD_a (%)", "V_THD_b (%)", "V_THD_c (%)"],
        "Current THD per phase", "I_THD (%)",
        ["I_THD_a (%)", "I_THD_b (%)", "I_THD_c (%)"],
    )
    add_two_chart_sheet(
        "Voltage & Current",
        "Per-phase L-N voltage", "V (V)",
        ["V_LN_a (V)", "V_LN_b (V)", "V_LN_c (V)"],
        "Per-phase current", "I (A)",
        ["I_a (A)", "I_b (A)", "I_c (A)"],
    )

    # Summary sheet — first
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum["A1"] = "Fluke 3540 FC Session Summary"
    ws_sum["A1"].font = Font(size=16, bold=True)
    ws_sum.merge_cells("A1:C1")
    if narrative:
        ws_sum["A2"] = "Executive summary: " + narrative
        ws_sum["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_sum.merge_cells("A2:F2")

    cfg = config or {}
    total = summary["sec_import"] + summary["sec_export"] + summary["sec_idle"]
    pct = lambda n: f"{(n / total * 100):.1f}%" if total else "n/a"

    summary_rows: list[tuple[str, str]] = []
    if cfg.get("asset_name"):
        summary_rows.append(("Asset", str(cfg["asset_name"])))
    if cfg.get("team_name"):
        summary_rows.append(("Team", str(cfg["team_name"])))
    if cfg.get("type") or cfg.get("firmware_version"):
        instr = " ".join(filter(None, [
            str(cfg.get("type") or ""),
            f"fw {cfg['firmware_version']}" if cfg.get("firmware_version") else "",
        ])).strip()
        if instr:
            summary_rows.append(("Instrument", instr))
    summary_rows.extend([
        ("Records (per-second)", f"{summary['rows']:,}"),
        ("", ""),
        ("Net energy (kWh)",         f"{summary['wh_sum'] / 1000:,.2f}"),
        ("Imported (kWh)",           f"{summary['wh_fwd'] / 1000:,.2f}"),
        ("Exported (kWh)",           f"{summary['wh_rev'] / 1000:,.2f}"),
        ("", ""),
        ("Peak import power (kW)",   f"{summary['p_peak_pos'] / 1000:,.2f}"),
        ("Peak export power (kW)",   f"{summary['p_peak_neg'] / 1000:,.2f}"),
        ("Peak current (A)",         f"{summary['i_peak']:.1f}"),
        ("", ""),
        ("Time importing",           f"{summary['sec_import']:,} s ({pct(summary['sec_import'])})"),
        ("Time exporting",           f"{summary['sec_export']:,} s ({pct(summary['sec_export'])})"),
        ("Time idle (|P|<10W)",      f"{summary['sec_idle']:,} s ({pct(summary['sec_idle'])})"),
    ])
    for i, (label, value) in enumerate(summary_rows, start=3):
        ws_sum.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws_sum.cell(row=i, column=2, value=value)
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 50

    if stats:
        _add_stats_sheet(wb, stats)
    if tod_rows:
        _add_tod_sheet(wb, tod_rows)

    wb.save(output_path)
    return output_path
